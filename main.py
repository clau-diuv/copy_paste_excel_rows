import openpyxl
from copy import copy

file_name = "excel_file.xlsx"

workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.worksheets[0]


def search_target(target, column):
    """
    search target on column column to insert rows before.
    :param target: target to search for
    :param column: index of column to serch for target
    :return: index of row where target was found
    """
    lista_col = list(worksheet.columns)[column - 1]  # column A - first column, first element of a list is at position 0
    for index in range(len(lista_col)):
        if lista_col[index].value == target:
            return index + 1


def copy_paste(star_row, end_row):
    """
    This function will copy the cells from range start_row to end_row and paste them before a specified row index
    :param star_row:
    :param end_row:
    :return:
    """
    # insert rows to paste the values
    index_row_before = search_target('copy_before', 1)
    worksheet.insert_rows(index_row_before, (end_row - star_row) + 1)

    for row_index in range(star_row, end_row + 1):
        for col_index in range(1, worksheet.max_column + 1):
            worksheet.cell(index_row_before, col_index).value = worksheet.cell(row_index, col_index).value
            if worksheet.cell(row_index, col_index).has_style:  # copy cell style
                worksheet.cell(index_row_before, col_index)._style = copy(worksheet.cell(row_index, col_index)._style)
        index_row_before += 1


def main():
    for i in range(3):  # copy rows for 3 times
        copy_paste(3, 6)
    workbook.save('after_inserting.xlsx')


main()
